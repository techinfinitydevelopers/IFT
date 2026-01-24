import openpyxl
import json

wb = openpyxl.load_workbook('Parameters_Sheet/Jury_Evaluation Sheet-2.xlsx')
ws = wb['Parameters with indicators ']

# Find the parameters - they appear to be in rows 3-17 based on structure
parameters = []

for row_idx in range(3, 18):
    row = [ws.cell(row=row_idx, column=col).value for col in range(1, 9)]
    param_name = row[1]
    
    if param_name and param_name.strip().replace('\xa0', '') not in ['Total Team Score', 'Grand Total', 'Comments', '']:
        # Get scale descriptions from columns 3-7 (ratings 1-5)
        scale_1 = ws.cell(row=row_idx, column=3).value
        scale_2 = ws.cell(row=row_idx, column=4).value
        scale_3 = ws.cell(row=row_idx, column=5).value
        scale_4 = ws.cell(row=row_idx, column=6).value
        scale_5 = ws.cell(row=row_idx, column=7).value
        
        param = {
            'name': param_name.strip().replace('\xa0', ''),
            'scale_1': str(scale_1).strip() if scale_1 else '',
            'scale_2': str(scale_2).strip() if scale_2 else '',
            'scale_3': str(scale_3).strip() if scale_3 else '',
            'scale_4': str(scale_4).strip() if scale_4 else '',
            'scale_5': str(scale_5).strip() if scale_5 else '',
        }
        parameters.append(param)

# Save to JSON file
with open('evaluation_rubric.json', 'w', encoding='utf-8') as f:
    json.dump(parameters, f, indent=2, ensure_ascii=False)

print(f"Total parameters found: {len(parameters)}")
print("\nParameter names:")
for i, p in enumerate(parameters, 1):
    print(f"{i}. {p['name']}")

print("\nSaved to evaluation_rubric.json")
