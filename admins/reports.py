"""Excel report generation helpers for the Super Admin Reports page."""
from django.http import HttpResponse

XLSX_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)


def xlsx_response(filename, headers, rows, sheet_title='Report'):
    """Build an .xlsx download response from headers + row tuples.

    - filename: download filename (without extension)
    - headers: list of column titles
    - rows: iterable of row sequences (values must be str/number/None)
    - Header row is bold + frozen; columns auto-sized (roughly).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or 'Report'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='5E2A97')
    center = Alignment(vertical='center')

    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.freeze_panes = 'A2'

    widths = [len(str(h)) for h in headers]
    for row in rows:
        vals = list(row)
        ws.append(vals)
        for i, v in enumerate(vals):
            if i < len(widths):
                widths[i] = max(widths[i], len(str(v)) if v is not None else 0)

    for i, w in enumerate(widths, start=1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 60)

    resp = HttpResponse(content_type=XLSX_CONTENT_TYPE)
    resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(resp)
    return resp
